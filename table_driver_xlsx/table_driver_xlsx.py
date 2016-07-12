import openpyxl


def value_generator(filename, sheet_index=None, header=True):
    """Generates cell index and value tuples for a sheet in an Excel file.

    Returns a generator that yields tuples of each cell that contains data
    for a worksheet (the specified sheet or active if None given). The form
    of the returned tuples is (row, column, value) where the row and column
    values are ints and the value is coerced to a str. Cells with no data
    are skipped. If 'header' is True, the first row is skipped and all
    subsequent rows have their row values decremented by one to compensate
    for the missing row.

    TODO:  Look into if the accepted type for the 'filename' parameter is
        IO[bytes] or IO[str] (open with 'r' or 'rb'?)

    Args:
        filename (str, IO[bytes]): Path to the file to open or file object.
        sheet_index (None, optional): Index for the sheet to generate values
            from. Default uses the active sheet. Defaults to None.
        header (bool, optional): Indicates if the sheet contains a header
            row. Defaults to True.

    Yields:
        Tuple(int, int, str): Data from a single cell of the Excel one at a
            time. Each tuple consists of the cell row and column, and the
            value as a string. Empty cells are not yielded.
    """
    # Load the workbook
    workbook = openpyxl.load_workbook(filename, read_only=True)

    # Select the worksheet
    worksheet = _get_sheet_from_workbook(workbook, sheet_index=sheet_index)

    yield _flat_cell_index_value_tuples(worksheet.iter_rows(), header=header)


def _get_sheet_from_workbook(workbook, sheet_index=None):
    """Returns a sheet from a Workbook; default active or at the given index

    By default returns the active worksheet from the given openpyxl.Workbook
    instance, but if given a sheet_index argument, the sheet at that index
    will instead be returned

    Args:
        workbook (openpyxl.Workbook): The workbook to select a sheet from.
        sheet_index (int, optional): The index of the worksheet to return.
            A value of None returns the active worksheet. Defaults to None.

    Returns:
        openpyxl.worksheet.Worksheet: The active or selected worksheet.
    """
    if sheet_index is None:
        return workbook.active
    return workbook.worksheets[sheet_index]


def _flat_cell_index_value_tuples(rows, header=True):
    """Flattens a 2D iterable of Cells to tuples with indicies and values.

    Loops through an iterable, of which each element should be and inner
    iterable object composed of openpyxl.cell.Cell instances. Returns
    a generator object that yields the cell row, column and value as a
    string (in that order) as a tuple. If Cell has no value, the entire
    cell is skipped and nothing is yielded for that instance.

    The intent is to pass this function the resulting generator returned
    from calling the openpyxl.sheet.Worksheet.iter_rows method, hence the
    naming convention of the parameter being rows and requiring all the
    innermost object to be openpyxl.cell.Cell instances. This also means
    that the outermost iterable passed to this function must be in
    row-major order (as the iter_rows method generates).

    If 'header' is True, we treat the first row of the 'rows' argument as
    being a header

    Args:
        rows (Iterable[Iterable[openpyxl.cell.Cell]]): Takes a 2D iterable
            of which the inner-most objects are openpyxl Cell instances.
        header (bool): Indicates if the first row in the Excel file was a
            row of headers.

    Yields:
        Tuple(int, int, str): Each cell returns a tuple consisting of the
            row and column indicies and the value as a string.
    """
    for row in rows:
        for cell in row:
            if cell.value is not None:
                yield (cell.row, cell.column, str(cell.value))
